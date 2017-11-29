"""grades.py generates an excel spreadsheet for you to enter your grades
and calculate your final average"""
from openpyxl import Workbook
from openpyxl.styles import Font, colors

class Category:
    """A category is represented by its name, the number of items in the category,
    and its weight on your final grade as a percentage"""
    def __init__(self, name, num, weight):
        self.name = name
        self.num = num
        self.weight = weight

# Record essentials such as the name of the course,
# the number of categories to generate, the weight of the category,
# and how many items are in each category

sheetName = raw_input("Name of Course: ")
print '\n'
numCategories = raw_input("Number of categories: ")
print '\n'
categories = list()
total = 0
for x in range(0, int(numCategories)):
    name = raw_input("Category " + str(x+1) + ": ")
    print '\n'
    num = raw_input("Number of elements " + ": ")
    print '\n'
    weight = raw_input("Weight of category (Ex: 15 = 15%) " + ": ")
    print '\n'
    categories.append(Category(name, num, weight))
    total += int(weight)
# All totals must equal 100 because they are percents
if ( int(total) != 100 ):
    raise ValueError('Your weights did not add up to 100')

# This is where we actually create the workbook and populate
# each category with the proper number of elements
wb = Workbook()
ws = wb.active
ws.title = sheetName
col = 1
j = int(1)
ws.cell(row=int(categories[0].num) + 4, column=2).value = '=SUM('

for i in range( 1, int(numCategories)+1):
    ws.cell( row=1, column=col ).value = categories[i-1].name
    for j in range(1, int(categories[i-1].num)+1):
        ws.cell(row=j+1, column=col).value = categories[i-1].name + " " + str(j)
    ws.cell(row=j+2, column=col).value = categories[i-1].name + " Average: "
    ws.cell(row=j+2, column=col+1).value = "=AVERAGE(" + ws.cell(row=2, column=col+1).coordinate + ":" + ws.cell(row=j+1, column=col+1).coordinate + ")"
    ws.cell(row=int(categories[0].num) + 4, column=2).value += ',' + ws.cell(row=j+2, column=col+1).coordinate + '*' + str(float(categories[i-1].weight)/100)
    col += 2
# Here we apply the finishing touches to highlight the final grade in red and
# make sure the formulas have their closing parenthesis
ft = Font(color=colors.RED)
ws.cell(row=int(categories[0].num) + 4, column=2).value += ')' 
ws.cell(row=int(categories[0].num) + 4, column=2).font = ft 
ws.cell(row=int(categories[0].num) + 4, column=1).value = 'Final Grade'
wb.save(sheetName + ".xlsx")
